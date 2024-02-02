#import win32com.client as win32 # Solo para convertir xls -> xlsx
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import pandas
import win32com.client
import win32api

import funcionesGeneral
import funcionesSQL
import os
from funcionesGeneral import buscarConfig
from datetime import datetime
from datetime import timedelta

def cargarHojasExcel(expediente):
    year = "20" + str(expediente[3:5])
    ruta = buscarConfig("rutas", "ruta_base_exp") + year + '/' + expediente + '/' + expediente + '.xlsx'
    if os.path.isfile(ruta) == False:
        return "nodisp"
    else:
        try:
            libro = load_workbook(filename=ruta)
            return libro.sheetnames
        except:
            return "error"

def guardarExcel(expediente):

    libro = Workbook()
    sheet = libro.active

    sheet["B1"] = "Número de expediente"
    sheet["B2"] = expediente

    libro.save(filename=expediente + ".xlsx")

    return True

def cargarDatosEnsayos(expediente):
    year = "20" + str(expediente[3:5])
    ruta = buscarConfig("rutas", "ruta_base_exp") + year + '/' + expediente + '/' + expediente + '.xlsx'
    if os.path.isfile(ruta) == False:
        return "nodisp"
    else:
        try:
            libro = load_workbook(filename=ruta, data_only=True)
            datos = {"sdv_fecha": [], "sdv_bajo_valor": [], "sdv_alto_valor": [], "sdv_alarma_tension_valor": [],
                     "ddr_h_fecha": [], "ddr_h_ancho_lob_valor": [], "ddr_h_aten_sec_valor": [], "ddr_h_desv_eje_valor": [], "ddr_h_frec_media_valor": [],
                     "ddr_v_fecha": [], "ddr_v_ancho_lob_valor": [], "ddr_v_aten_sec_valor": [], "ddr_v_desv_eje_valor": [], "ddr_v_frec_media_valor": [],
                     "ddr_frec_anterior": "",
                     "tr_e_fecha": [], "tr_e_bajo_valor": [], "tr_e_alto_valor": [],
                     "tr_m_fecha": [], "tr_m_bajo_valor": [], "tr_m_alto_valor": []}

            for hoja in libro.sheetnames:
                if str("alej").upper() in hoja.upper() or str("aprox").upper() in hoja.upper():
                    datos["sdv_bajo_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_sdv", "sdv_bajo")].value, 2))
                    datos["sdv_alto_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_sdv", "sdv_alto")].value, 2))
                    datos["sdv_fecha"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_sdv", "sdv_fecha")].value)
                    if libro[hoja][funcionesGeneral.buscarConfig("celdas_sdv", "sdv_alarma_tension")].value is not None:
                        datos["sdv_alarma_tension_valor"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_sdv", "sdv_alarma_tension")].value)
                elif str("traf").upper() in hoja.upper() and str("_e").upper() in hoja.upper():
                    datos["tr_e_bajo_valor"].append (round(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_bajo")].value, 1))
                    datos["tr_e_alto_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_alto")].value, 1))
                    datos["tr_e_fecha"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_fecha")].value)
                elif str("traf").upper() in hoja.upper() and str("_m").upper() in hoja.upper():
                    datos["tr_m_bajo_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_bajo")].value, 1))
                    datos["tr_m_alto_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_alto")].value, 1))
                    datos["tr_m_fecha"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_trafreal", "tr_fecha")].value)
                elif str("ddr_h").upper() in hoja.upper() and len(hoja) <= 5:
                    datos["ddr_h_fecha"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_fecha")].value)
                    # datos["ddr_h_ancho_lob_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_ancho_lob")].value, 2))
                    # datos["ddr_h_aten_sec_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_aten_sec")].value, 2))
                    # datos["ddr_h_desv_eje_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_desv_eje")].value, 2))
                    # datos["ddr_h_frec_media_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_media")].value))
                    if libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_anterior")].value is not None:
                        datos["ddr_frec_anterior"] = str(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_anterior")].value, 3)).replace(".",",")
                elif str("ddr_h").upper() in hoja.upper() and len(hoja) > 5: # Hoja Datos_DdR_H
                    datos["ddr_h_ancho_lob_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_ancho_lob")].value))
                    datos["ddr_h_aten_sec_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_aten_sec")].value))
                    datos["ddr_h_desv_eje_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_desv_eje")].value, 2))
                    datos["ddr_h_frec_media_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_media")].value, 3))
                elif str("ddr_v").upper() in hoja.upper() and len(hoja) <= 5:
                    datos["ddr_v_fecha"].append(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_fecha")].value)
                    # datos["ddr_v_ancho_lob_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_ancho_lob")].value, 2))
                    # datos["ddr_v_aten_sec_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_aten_sec")].value, 2))
                    # datos["ddr_v_desv_eje_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_desv_eje")].value, 2))
                    # datos["ddr_v_frec_media_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_media")].value, 3))
                elif str("ddr_v").upper() in hoja.upper() and len(hoja) > 5: # Hoja Datos_DdR_V
                    datos["ddr_v_ancho_lob_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_ancho_lob")].value))
                    datos["ddr_v_aten_sec_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_aten_sec")].value))
                    datos["ddr_v_desv_eje_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_desv_eje")].value, 2))
                    datos["ddr_v_frec_media_valor"].append(round(libro[hoja][funcionesGeneral.buscarConfig("celdas_ddr", "ddr_frec_media")].value, 3))

            return datos
        except:
            print(hoja)
            return "error"

def generarPdf(df, numExp, numColAdicional, precintos):

    ruta = funcionesGeneral.buscarConfig("rutas", "ruta_plantilla_resultados")

    if os.path.isfile(ruta) == False:
        return "nodisp"
    else:
        try:
            libro = load_workbook(filename=ruta)
            hoja = libro.active

            numLineasMax = 0
            for i in range(len(df)):
                for j in range(numColAdicional):
                    if df.iat[i,3+j]: # Da igual buscar así (en columnas "Campo"), que sumar numColAdicional y buscar en columnas "Valor"
                        if "texto_columna" in df.iat[i,3+j]:
                            if df.iat[i, 3+j+numColAdicional] == "": # Ocultar filas de texto si los valores están vacíos
                                hoja.row_dimensions[int(df.iat[i,0])].hidden = True
                            else:
                                numLineas = len(str(df.iat[i, 3+j+numColAdicional]).splitlines())
                                #alturaOrig = hoja.row_dimensions[int(df.iat[i,0])].height
                                if numLineas > numLineasMax:
                                    numLineasMax = numLineas
                                hoja.row_dimensions[int(df.iat[i,0])].height = 9*numLineasMax

                        for columna in range(1,19):
                            if str(hoja.cell(int(df.iat[i, 0]), columna).value) == str(j+1): # df.iat[i, 0] es donde está el número de fila en el df
                                #hoja.cell(int(df.iat[i, 0]), columna).alignment = Alignment(wrapText=True)
                                hoja.cell(int(df.iat[i, 0]), columna).value = str(df.iat[i, 3+j+numColAdicional])

            # Se recorren todas las filas de la plantilla y se comparan con el df para ver si el número de fila está contenido en él (si no, se oculta)
            for i in range(int(funcionesGeneral.buscarConfig("celdas_certificado", "inicio_certif")), int(funcionesGeneral.buscarConfig("celdas_certificado", "final_certif"))+1):
                if str(i) not in df["Fila"].values:
                    hoja.row_dimensions[i].hidden = True
            
            if precintos:
                hoja[funcionesGeneral.buscarConfig("celdas_certificado", "celda_precintos")].value =  'Precintos:\n' + '\n'.join(str(p) for p in precintos)

            year = "20" + str(numExp[3:5])
            ruta = buscarConfig("rutas", "ruta_base_exp") + year + '/' + numExp + '/'

            libro.save(ruta + 'resultados_temp.xlsx')

            excel = win32com.client.Dispatch('Excel.Application')
            libro_excel = excel.Workbooks.Open(os.path.abspath(ruta + 'resultados_temp.xlsx'))
            hoja_result = libro_excel.Sheets('Resultados')
            rango_result = hoja_result.Range('B' + funcionesGeneral.buscarConfig("celdas_certificado", "inicio_certif") + 
                                             ':P' + funcionesGeneral.buscarConfig("celdas_certificado", "final_certif"))

            rango_result.ExportAsFixedFormat(0, os.path.abspath(ruta + numExp + '.pdf'))
            libro_excel.Close(SaveChanges=False)
            os.remove(os.path.abspath(ruta + 'resultados_temp.xlsx'))
            # excel.Quit()

            return "ok"
        except:
            return "error"

def convertirEnsayos(expediente):
    year = "20" + str(expediente[3:5])
    ruta = buscarConfig("rutas", "ruta_base_exp") + year + '/' + expediente

    excel = win32com.client.Dispatch('Excel.Application')
    libroNuevo = excel.Workbooks.Add()
    libroNuevo.SaveAs(ruta + "/" + expediente + ".xlsx")
    excel.Application.Visible = False
    excel.Application.ScreenUpdating = False
    excel.Application.AskToUpdateLinks = False
    excel.Application.DisplayAlerts = False
    for archivo in os.listdir(os.path.abspath(ruta)):
        if os.path.splitext(archivo)[1] == ".XLS" or os.path.splitext(archivo)[1] == ".xls":
            print(os.path.splitext(archivo)[0])
            if os.path.splitext(archivo)[0] == expediente + "a":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123")
                libro.Sheets(1).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="Aproximacion" 
                libro.Close(savechanges:=False)
            elif os.path.splitext(archivo)[0] == expediente + "d":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123")
                libro.Sheets(1).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="Alejamiento" 
                libro.Close(savechanges:=False)
            elif os.path.splitext(archivo)[0] == expediente + "h":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123", WriteResPassword:="cinem123")
                libro.Sheets(2).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="DdR_H"
                libro.Sheets(1).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="Datos_DdR_H"  
                libro.Close(savechanges:=False)
            elif os.path.splitext(archivo)[0] == expediente + "v":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123", WriteResPassword:="cinem123")
                libro.Sheets(2).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="DdR_V"
                libro.Sheets(1).Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="Datos_DdR_V"  
                libro.Close(savechanges:=False)
            elif os.path.splitext(archivo)[0] == expediente + "e":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123", WriteResPassword:="cinem123", UpdateLinks:=False)
                excel.Application.DisplayAlerts = False
                if libro.Sheets("InformeGrabDoppler").Visible == -1:    # Visible (no visible sería 0)
                    libro.Sheets("InformeGrabDoppler").Copy(Before:=libroNuevo.Sheets(1))
                if libro.Sheets("Informe").Visible == -1:
                    libro.Sheets("Informe").Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="TraficoReal_e" 
                libro.Close(savechanges:=False)
            elif os.path.splitext(archivo)[0] == expediente + "m":
                rutaCompleta = ruta + '/' + archivo
                libro = excel.Workbooks.Open(rutaCompleta, False, False, None, password:="cinem123", WriteResPassword:="cinem123", UpdateLinks:=False)
                excel.Application.DisplayAlerts = False
                if libro.Sheets("InformeGrabDoppler").Visible == -1:
                    libro.Sheets("InformeGrabDoppler").Copy(Before:=libroNuevo.Sheets(1))
                if libro.Sheets("Informe").Visible == -1:
                    libro.Sheets("Informe").Copy(Before:=libroNuevo.Sheets(1))
                libroNuevo.Sheets(1).Name="TraficoReal_m" 
                libro.Close(savechanges:=False)
    libroNuevo.Application.DisplayAlerts = False
    excel.Application.ScreenUpdating = True
    excel.Application.Visible = True
    libroNuevo.Sheets("Hoja1").Delete
    libroNuevo.Close(savechanges:=True)

def copiarHojaLibros(rutaLibro, nombreNuevoHoja, rutaPlantilla, nombrePlantilla, tipoSelect):
    contador_tr = 0
    if os.path.isfile(rutaLibro) == False:
        wb_nuevo = Workbook()
    else:
        wb_nuevo = load_workbook(filename=rutaLibro, data_only=True)
        for hoja in wb_nuevo.sheetnames:
            if str(nombreNuevoHoja).upper() in hoja.upper():
                contador_tr += 1
    wb_nuevo.save(rutaLibro)
    wb_nuevo.close()

    if contador_tr > 0:
        contador_tr = "_" + str(contador_tr)
    else:
        contador_tr = ""

    # Única forma de copiar hojas entre libros distintos
    excel = win32com.client.Dispatch('Excel.Application')
    excel_nuevo = excel.Workbooks.Open(os.path.abspath(rutaLibro))
    excel_plantilla = excel.Workbooks.Open(os.path.abspath(rutaPlantilla))

    # Código para copiar cabeceras
    hoja_plantilla_config = excel_plantilla.Sheets("Configuracion")
    hoja_plantilla_informe = excel_plantilla.Sheets(nombrePlantilla)

    if tipoSelect == "cinemCab": hoja_plantilla_config.Range("B3:N5").Copy()
    elif tipoSelect == "cinemCab": hoja_plantilla_config.Range("B7:N9").Copy()
    elif tipoSelect ==  "cinemEst" or tipoSelect == "cinemMov": hoja_plantilla_config.Range("B11:N13").Copy()

    hoja_plantilla_informe.Paste(hoja_plantilla_informe.Range("B6:N8"))

    excel.Application.CutCopyMode = False


    # excel = win32com.client.Dispatch('Excel.Application')
    #         libro_excel = excel.Workbooks.Open(os.path.abspath(ruta + 'resultados_temp.xlsx'))
    #         hoja_result = libro_excel.Sheets('Resultados')
    #         rango_result = hoja_result.Range('B' + funcionesGeneral.buscarConfig("celdas_certificado", "inicio_certif") + 
    #                                          ':P' + funcionesGeneral.buscarConfig("celdas_certificado", "final_certif"))

    #         rango_result.ExportAsFixedFormat(0, os.path.abspath(ruta + numExp + '.pdf'))
    #         libro_excel.Close(SaveChanges=False)



    excel_plantilla.Sheets(nombrePlantilla).Copy(After:=excel_nuevo.Sheets(excel_nuevo.Sheets.Count))
    nombreNuevoHoja = nombreNuevoHoja + contador_tr
    excel_nuevo.Sheets(nombrePlantilla).Name = nombreNuevoHoja
    excel_nuevo.Close(SaveChanges=True)
    excel_plantilla.Close(SaveChanges=False)

    wb_nuevo = load_workbook(filename=rutaLibro, data_only=True)
    for nombreHoja in wb_nuevo.sheetnames:
        if nombreHoja == 'Sheet':
            hoja = wb_nuevo.get_sheet_by_name('Sheet')
            wb_nuevo.remove_sheet(hoja)
            wb_nuevo.save(rutaLibro)
    wb_nuevo.close
    return nombreNuevoHoja

def guardarTrafReal(df, expediente, tipoSelect):
    # Las columnas 5 y 6 contienen los errores, pero los campos que están vacíos son de tipo string
    # por lo que dan error al calcular abs
    lista_errorKmh = []
    lista_errorPorcent = []
    for i in range(len(df)):
        if df.iat[i,5] and df.iat[i,5] != "":
            lista_errorKmh.append(df.iat[i,5])
        if df.iat[i,6] and df.iat[i,6] != "":
            lista_errorPorcent.append(df.iat[i,6])

    year = "20" + str(expediente[3:5])
    ruta = buscarConfig("rutas", "ruta_base_exp") + year + '/' + expediente + '/' + expediente + '.xlsx'

    if tipoSelect == 'cinemMov': nombreHoja_tr = copiarHojaLibros(ruta, "Trafico_real_m", buscarConfig("rutas", "ruta_plantilla_trafreal"), "Plantilla_TraficoReal", tipoSelect)
    else: nombreHoja_tr = copiarHojaLibros(ruta, "Trafico_real_e", buscarConfig("rutas", "ruta_plantilla_trafreal"), "Plantilla_TraficoReal", tipoSelect)
   
    wb_nuevo = load_workbook(filename=ruta, data_only=True)
    hoja = wb_nuevo[nombreHoja_tr]

    for i in range(len(df)):
        wb_nuevo[nombreHoja_tr]["D" + str(i+11)] = df.iat[i,0]
        wb_nuevo[nombreHoja_tr]["E" + str(i+11)] = df.iat[i,1]
        wb_nuevo[nombreHoja_tr]["F" + str(i+11)] = df.iat[i,2]
        wb_nuevo[nombreHoja_tr]["G" + str(i+11)] = df.iat[i,3]
        wb_nuevo[nombreHoja_tr]["H" + str(i+11)] = df.iat[i,4]
        wb_nuevo[nombreHoja_tr]["J" + str(i+11)] = df.iat[i,5]
        wb_nuevo[nombreHoja_tr]["L" + str(i+11)] = df.iat[i,6]
        wb_nuevo[nombreHoja_tr]["P" + str(i+11)] = df.iat[i,7]

    if max(lista_errorKmh) >= abs(min(lista_errorKmh)):
        wb_nuevo[nombreHoja_tr]["L63"] = max(lista_errorKmh)
    else:
        wb_nuevo[nombreHoja_tr]["L63"] = min(lista_errorKmh)

    if max(lista_errorPorcent) >= abs(min(lista_errorPorcent)):
        wb_nuevo[nombreHoja_tr]["L64"] = max(lista_errorPorcent)
    else:
        wb_nuevo[nombreHoja_tr]["L64"] = min(lista_errorPorcent)

    usuario = ""
    for palabra in win32api.GetUserNameEx(3).split():
        usuario = usuario + palabra[0]
    wb_nuevo[nombreHoja_tr]["J66"] = usuario
    wb_nuevo[nombreHoja_tr]["C68"] = expediente
    wb_nuevo[nombreHoja_tr]["H70"] = (datetime.now() - timedelta(minutes=5)).strftime("%H:%M") +  " - " + datetime.now().strftime("%H:%M")
    wb_nuevo[nombreHoja_tr]["J70"] = datetime.now().today().strftime("%d/%m/%Y")

    wb_nuevo.save(ruta)
        
#Éste no hará falta
def tablaDesdeExcel():
    pd_Excel = pandas.read_excel('plantillas/matrices.xlsx')
    funcionesSQL.crearSQLExcel(pd_Excel)